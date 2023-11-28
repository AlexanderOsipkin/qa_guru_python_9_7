import zipfile
from openpyxl import load_workbook


def test_text_in_xlsx_file():
    with zipfile.ZipFile('resources/archive.zip') as zip_file:
        with zip_file.open('tmp/file_xlsx.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            first_title = sheet.cell(row=1, column=1).value
            second_title = sheet.cell(row=1, column=2).value
            third_title = sheet.cell(row=1, column=3).value
            assert first_title == "Внешний идентификатор для импорта"
            assert second_title == "Вышестоящий отдел"
            assert third_title == "Название"
